"""报表业务逻辑"""
import io
from datetime import date
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from app.extensions import db
from app.models.attendance import AttendanceTask, AttendanceRecord
from app.models.student import Student, StudentCourse
from app.models.course import Course


class ReportService:
    def get_course_stats(self, course_id: int, date_from: str, date_to: str) -> dict:
        q = AttendanceTask.query.filter_by(course_id=course_id, status="finished")
        if date_from:
            q = q.filter(AttendanceTask.task_date >= date.fromisoformat(date_from))
        if date_to:
            q = q.filter(AttendanceTask.task_date <= date.fromisoformat(date_to))
        tasks = q.order_by(AttendanceTask.task_date).all()
        return {
            "course_id": course_id,
            "series": [
                {
                    "date": t.task_date.isoformat(),
                    "task_id": t.id,
                    "total": t.total_students,
                    "present": t.present_count,
                    "rate": t.get_attendance_rate(),
                }
                for t in tasks
            ],
        }

    def get_student_history(self, student_id: int) -> dict:
        records = (
            db.session.query(AttendanceRecord, AttendanceTask, Course)
            .join(AttendanceTask, AttendanceRecord.task_id == AttendanceTask.id)
            .join(Course, AttendanceTask.course_id == Course.id)
            .filter(AttendanceRecord.student_id == student_id)
            .order_by(AttendanceTask.task_date.desc())
            .all()
        )
        total = len(records)
        present = sum(1 for r, _, _ in records if r.status == "present")
        return {
            "student_id": student_id,
            "total_tasks": total,
            "present_count": present,
            "attendance_rate": round(present / total, 4) if total else 0.0,
            "records": [
                {
                    **rec.to_dict(),
                    "task_date": task.task_date.isoformat(),
                    "course_name": course.course_name,
                }
                for rec, task, course in records
            ],
        }

    def export_excel(self, filters: dict) -> bytes:
        wb = openpyxl.Workbook()
        # Sheet 1: 考勤汇总
        ws1 = wb.active
        ws1.title = "考勤汇总"
        header = ["学号", "姓名", "班级", "出勤次数", "缺勤次数", "待核实次数", "出勤率"]
        self._write_header(ws1, header)
        students = Student.query.order_by(Student.student_id).all()
        for row_idx, st in enumerate(students, start=2):
            recs = AttendanceRecord.query.filter_by(student_id=st.id).all()
            present = sum(1 for r in recs if r.status == "present")
            absent = sum(1 for r in recs if r.status == "absent")
            unverified = sum(1 for r in recs if r.status == "unverified")
            total = len(recs)
            rate = f"{present / total:.1%}" if total else "N/A"
            ws1.append([st.student_id, st.name, st.class_name,
                        present, absent, unverified, rate])
        # Sheet 2: 考勤明细
        ws2 = wb.create_sheet("考勤明细")
        detail_header = ["课程名", "考勤日期", "学号", "姓名", "状态", "置信度", "识别时间"]
        self._write_header(ws2, detail_header)
        rows = (
            db.session.query(AttendanceRecord, AttendanceTask, Course, Student)
            .join(AttendanceTask, AttendanceRecord.task_id == AttendanceTask.id)
            .join(Course, AttendanceTask.course_id == Course.id)
            .join(Student, AttendanceRecord.student_id == Student.id)
            .order_by(AttendanceTask.task_date.desc())
            .all()
        )
        for rec, task, course, st in rows:
            ws2.append([
                course.course_name,
                task.task_date.isoformat(),
                st.student_id,
                st.name,
                rec.status,
                round(rec.confidence, 3) if rec.confidence else "",
                rec.recognized_at.strftime("%Y-%m-%d %H:%M:%S") if rec.recognized_at else "",
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf.read()

    def _write_header(self, ws, headers: list):
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center")

    def get_warnings(self) -> list:
        from sqlalchemy import func, text
        # 查询连续缺勤>=3次的学生（简化版：近30条记录中缺勤>=3）
        subq = (
            db.session.query(
                AttendanceRecord.student_id,
                func.sum(
                    db.case((AttendanceRecord.status == "absent", 1), else_=0)
                ).label("absent_count"),
                func.count(AttendanceRecord.id).label("total"),
            )
            .group_by(AttendanceRecord.student_id)
            .having(func.sum(
                db.case((AttendanceRecord.status == "absent", 1), else_=0)
            ) >= 3)
            .subquery()
        )
        rows = db.session.query(Student, subq.c.absent_count).join(
            subq, Student.id == subq.c.student_id
        ).all()
        return [
            {
                "student_id": st.student_id,
                "name": st.name,
                "class_name": st.class_name,
                "absent_count": absent,
            }
            for st, absent in rows
        ]

    def get_dashboard_stats(self) -> dict:
        """仪表盘快速统计：学生总数 / 今日任务 / 今日出勤率 / 待核实记录"""
        from datetime import date as _date
        from sqlalchemy import func
        student_count = db.session.query(func.count(Student.id)).scalar() or 0
        today = _date.today()
        today_tasks = AttendanceTask.query.filter_by(task_date=today).all()
        today_task_count = len(today_tasks)
        total_s = sum(t.total_students or 0 for t in today_tasks)
        present_s = sum(t.present_count or 0 for t in today_tasks)
        today_rate = round(present_s / total_s, 4) if total_s else 0.0
        unverified_count = AttendanceRecord.query.filter_by(
            status="unverified"
        ).count()
        return {
            "student_count": student_count,
            "today_task_count": today_task_count,
            "today_attendance_rate": today_rate,
            "unverified_count": unverified_count,
        }

    def check_and_generate_warnings(self) -> int:
        """定时任务调用：返回新增预警数"""
        warnings = self.get_warnings()
        return len(warnings)
